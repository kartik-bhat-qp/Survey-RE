#!/usr/bin/env python3
"""Consolidate transcript CSV files into one Excel workbook grouped by question.

Reads one or more transcript CSVs, groups every response under the question it
answers (matching questions that are worded slightly differently across files),
and writes an .xlsx workbook with a summary sheet, a combined sheet, and one
sheet per question.

Usage:
    python3 transcripts_to_excel.py --output out.xlsx interviews.csv focus-group.csv
    python3 transcripts_to_excel.py --output out.xlsx --print-summary *.csv

Three CSV layouts are detected automatically:
    long      question/answer columns, one row per response
    wide      one row per respondent, one column per question
    dialogue  speaker/text columns (Zoom, Otter, Teams style exports)
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from collections import Counter, OrderedDict
from dataclasses import dataclass, field
from datetime import datetime
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - dependency guard
    sys.stderr.write(
        "openpyxl is required. Install it with:\n"
        "  python3 -m pip install -r requirements.txt\n"
    )
    raise SystemExit(2)

# Excel hard limits.
MAX_CELL_CHARS = 32_000
MAX_SHEET_NAME = 31

QUESTION_HEADERS = {
    "question",
    "questions",
    "question_text",
    "question_asked",
    "question_label",
    "prompt",
    "q",
    "ques",
    "item",
    "topic",
}
ANSWER_HEADERS = {
    "answer",
    "answers",
    "answer_text",
    "response",
    "responses",
    "response_text",
    "reply",
    "verbatim",
    "value",
}
SPEAKER_HEADERS = {
    "speaker",
    "speaker_name",
    "speaker_label",
    "speaker_id",
    "role",
    "who",
    "name",
}
TEXT_HEADERS = {
    "text",
    "utterance",
    "dialogue",
    "sentence",
    "content",
    "caption",
    "line",
    "message",
    "transcript_text",
}
TRANSCRIPT_HEADERS = {
    "transcript",
    "transcript_id",
    "transcript_name",
    "interview",
    "interview_id",
    "respondent",
    "respondent_id",
    "respondent_name",
    "participant",
    "participant_id",
    "conversation_id",
    "session",
    "session_id",
    "meeting",
    "file",
    "filename",
    "source",
}
# Columns that are never questions in a wide export.
METADATA_HEADERS = {
    "date",
    "time",
    "timestamp",
    "start",
    "start_time",
    "end",
    "end_time",
    "duration",
    "language",
    "email",
    "ip_address",
    "ip",
    "status",
    "progress",
    "finished",
    "created_at",
    "updated_at",
    "owner",
}

# Leading numbering to drop when matching questions: "Q1.", "3)", "[Q12]", "Q:".
QUESTION_PREFIX_RE = re.compile(r"^\s*[\[\(]?\s*(?:q(?:uestion)?\s*)?\d*\s*[\]\).:\-–]*\s*", re.I)
# Conversational lead-ins that would otherwise stop the same question matching.
LEAD_IN_RE = re.compile(
    r"^(?:(?:ok(?:ay)?|alright|all right|right|so|and|but|now|next|then|great|perfect"
    r"|cool|got it|makes sense|understood|thanks|thank you|awesome|interesting|sure"
    r"|yeah|yes|well|um|uh|to start|to begin|first|firstly|second|secondly|last"
    r"|lastly|finally|moving on|one more thing|one last thing|just curious"
    r"|out of curiosity|let me ask|i wanted to ask|can i ask)\b[\s,.:;–—-]+)+",
    re.I,
)
SENTENCE_SPLIT_RE = re.compile(r"(?<=[.!?])\s+")
NON_WORD_RE = re.compile(r"[^a-z0-9 ]+")
WHITESPACE_RE = re.compile(r"\s+")
INVALID_SHEET_CHARS_RE = re.compile(r"[\[\]:*?/\\]")


@dataclass
class Response:
    transcript: str
    speaker: str
    answer: str
    source_file: str


@dataclass
class QuestionGroup:
    key: str
    label: str
    responses: List[Response] = field(default_factory=list)
    label_variants: Counter = field(default_factory=Counter)
    unanswered: int = 0

    @property
    def transcripts(self) -> List[str]:
        seen: "OrderedDict[str, None]" = OrderedDict()
        for response in self.responses:
            seen.setdefault(response.transcript, None)
        return list(seen.keys())


def clean_text(value: Optional[str]) -> str:
    if value is None:
        return ""
    return WHITESPACE_RE.sub(" ", str(value).replace("\u00a0", " ")).strip()


def normalize_header(value: Optional[str]) -> str:
    slug = NON_WORD_RE.sub(" ", clean_text(value).lower())
    return WHITESPACE_RE.sub("_", slug).strip("_")


def extract_question(text: str) -> str:
    """Pull the asked question out of a spoken utterance.

    "Got it. How often do you refresh?" -> "How often do you refresh?"
    "To start, what tools do you use?" -> "What tools do you use?"
    """
    cleaned = clean_text(text)
    sentences = [part for part in SENTENCE_SPLIT_RE.split(cleaned) if part.strip()]
    asking = [part for part in sentences if "?" in part]
    if asking:
        cleaned = asking[-1].strip()
    cleaned = LEAD_IN_RE.sub("", cleaned).strip()
    if not cleaned:
        return clean_text(text)
    return cleaned[0].upper() + cleaned[1:]


def question_key(label: str) -> str:
    """Collapse wording differences so the same question matches across files."""
    text = QUESTION_PREFIX_RE.sub("", clean_text(label).lower())
    text = LEAD_IN_RE.sub("", text)
    text = NON_WORD_RE.sub(" ", text)
    return WHITESPACE_RE.sub(" ", text).strip()


def truncate(value: str, limit: int = MAX_CELL_CHARS) -> str:
    if len(value) <= limit:
        return value
    return value[: limit - 1] + "…"


def find_column(headers: Sequence[str], candidates: Iterable[str]) -> Optional[str]:
    """Exact header match first, then a contains match, preserving CSV order."""
    candidate_set = set(candidates)
    for header in headers:
        if normalize_header(header) in candidate_set:
            return header
    for header in headers:
        normalized = normalize_header(header)
        for candidate in candidate_set:
            if candidate in normalized.split("_"):
                return header
    return None


def read_rows(path: str) -> Tuple[List[str], List[Dict[str, str]]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(8192)
        handle.seek(0)
        try:
            dialect: "csv.Dialect | type[csv.Dialect]" = csv.Sniffer().sniff(
                sample, delimiters=",;\t|"
            )
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(handle, dialect=dialect)
        headers = [header for header in (reader.fieldnames or []) if header is not None]
        rows = [row for row in reader if any(clean_text(cell) for cell in row.values())]
    return headers, rows


def detect_layout(headers: Sequence[str]) -> Tuple[str, Dict[str, Optional[str]]]:
    question_col = find_column(headers, QUESTION_HEADERS)
    answer_col = find_column(headers, ANSWER_HEADERS)
    transcript_col = find_column(headers, TRANSCRIPT_HEADERS)
    speaker_col = find_column(headers, SPEAKER_HEADERS)
    text_col = find_column(headers, TEXT_HEADERS)

    columns = {
        "question": question_col,
        "answer": answer_col,
        "transcript": transcript_col,
        "speaker": speaker_col,
        "text": text_col,
    }

    if question_col and answer_col:
        return "long", columns
    if speaker_col and text_col:
        return "dialogue", columns
    return "wide", columns


def guess_interviewer(rows: List[Dict[str, str]], speaker_col: str, text_col: str) -> Optional[str]:
    """The speaker who asks the most questions is treated as the interviewer."""
    asked: Counter = Counter()
    for row in rows:
        text = clean_text(row.get(text_col))
        if "?" in text:
            asked[clean_text(row.get(speaker_col)) or "Unknown"] += 1
    if not asked:
        return None
    return asked.most_common(1)[0][0]


def collect_long(
    rows: List[Dict[str, str]],
    columns: Dict[str, Optional[str]],
    default_transcript: str,
    source_file: str,
) -> List[Tuple[str, Response, bool]]:
    question_col = columns["question"]
    answer_col = columns["answer"]
    transcript_col = columns["transcript"]
    speaker_col = columns["speaker"]
    collected: List[Tuple[str, Response, bool]] = []

    for row in rows:
        label = clean_text(row.get(question_col)) if question_col else ""
        if not label:
            continue
        answer = clean_text(row.get(answer_col)) if answer_col else ""
        transcript = (
            clean_text(row.get(transcript_col)) if transcript_col else ""
        ) or default_transcript
        speaker = clean_text(row.get(speaker_col)) if speaker_col else ""
        collected.append(
            (
                label,
                Response(
                    transcript=transcript,
                    speaker=speaker,
                    answer=answer,
                    source_file=source_file,
                ),
                bool(answer),
            )
        )
    return collected


def collect_wide(
    headers: Sequence[str],
    rows: List[Dict[str, str]],
    columns: Dict[str, Optional[str]],
    default_transcript: str,
    source_file: str,
) -> List[Tuple[str, Response, bool]]:
    transcript_col = columns["transcript"]
    speaker_col = columns["speaker"]
    skip = {transcript_col, speaker_col}
    question_cols = [
        header
        for header in headers
        if header not in skip and normalize_header(header) not in METADATA_HEADERS
    ]
    collected: List[Tuple[str, Response, bool]] = []

    for index, row in enumerate(rows, start=1):
        transcript = (
            clean_text(row.get(transcript_col)) if transcript_col else ""
        ) or f"{default_transcript} — row {index}"
        speaker = clean_text(row.get(speaker_col)) if speaker_col else ""
        for header in question_cols:
            label = clean_text(header)
            if not label:
                continue
            answer = clean_text(row.get(header))
            collected.append(
                (
                    label,
                    Response(
                        transcript=transcript,
                        speaker=speaker,
                        answer=answer,
                        source_file=source_file,
                    ),
                    bool(answer),
                )
            )
    return collected


def collect_dialogue(
    rows: List[Dict[str, str]],
    columns: Dict[str, Optional[str]],
    default_transcript: str,
    source_file: str,
) -> List[Tuple[str, Response, bool]]:
    speaker_col = columns["speaker"]
    text_col = columns["text"]
    transcript_col = columns["transcript"]
    interviewer = guess_interviewer(rows, speaker_col, text_col)
    collected: List[Tuple[str, Response, bool]] = []

    current_label: Optional[str] = None
    # Answers are accumulated per speaker so a follow-up stays with its author.
    pending: "OrderedDict[str, List[str]]" = OrderedDict()
    current_transcript = default_transcript

    def flush() -> None:
        if current_label is None:
            return
        if not pending:
            collected.append(
                (
                    current_label,
                    Response(
                        transcript=current_transcript,
                        speaker="",
                        answer="",
                        source_file=source_file,
                    ),
                    False,
                )
            )
            return
        for speaker, parts in pending.items():
            answer = clean_text(" ".join(parts))
            collected.append(
                (
                    current_label,
                    Response(
                        transcript=current_transcript,
                        speaker=speaker,
                        answer=answer,
                        source_file=source_file,
                    ),
                    bool(answer),
                )
            )

    for row in rows:
        text = clean_text(row.get(text_col))
        if not text:
            continue
        speaker = clean_text(row.get(speaker_col)) or "Unknown"
        transcript = (
            clean_text(row.get(transcript_col)) if transcript_col else ""
        ) or default_transcript

        is_question = "?" in text and (interviewer is None or speaker == interviewer)
        if is_question:
            flush()
            current_label = extract_question(text)
            current_transcript = transcript
            pending = OrderedDict()
            continue

        if current_label is None:
            continue
        if interviewer is not None and speaker == interviewer:
            # Interviewer commentary between answers is not a response.
            continue
        pending.setdefault(speaker, []).append(text)

    flush()
    return collected


def build_groups(paths: Sequence[str]) -> Tuple["OrderedDict[str, QuestionGroup]", List[Dict[str, object]]]:
    groups: "OrderedDict[str, QuestionGroup]" = OrderedDict()
    file_reports: List[Dict[str, object]] = []

    for path in paths:
        source_file = os.path.basename(path)
        default_transcript = os.path.splitext(source_file)[0]
        headers, rows = read_rows(path)
        if not headers:
            file_reports.append(
                {"file": source_file, "layout": "empty", "rows": 0, "responses": 0}
            )
            continue

        layout, columns = detect_layout(headers)
        if layout == "long":
            collected = collect_long(rows, columns, default_transcript, source_file)
        elif layout == "dialogue":
            collected = collect_dialogue(rows, columns, default_transcript, source_file)
        else:
            collected = collect_wide(headers, rows, columns, default_transcript, source_file)

        answered = 0
        for label, response, has_answer in collected:
            key = question_key(label)
            if not key:
                continue
            group = groups.get(key)
            if group is None:
                group = QuestionGroup(key=key, label=clean_text(label))
                groups[key] = group
            group.label_variants[clean_text(label)] += 1
            if has_answer:
                group.responses.append(response)
                answered += 1
            else:
                group.unanswered += 1

        file_reports.append(
            {
                "file": source_file,
                "layout": layout,
                "rows": len(rows),
                "responses": answered,
            }
        )

    for group in groups.values():
        if group.label_variants:
            group.label = group.label_variants.most_common(1)[0][0]

    return groups, file_reports


HEADER_FILL = PatternFill("solid", fgColor="F1F5F9")
QUESTION_FILL = PatternFill("solid", fgColor="E8F0FE")
TITLE_FONT = Font(bold=True, size=14, color="1F2937")
HEADER_FONT = Font(bold=True, color="374151")
QUESTION_FONT = Font(bold=True, size=11, color="1D4ED8")
THIN_BORDER = Border(bottom=Side(style="thin", color="E5E7EB"))
TOP_WRAP = Alignment(vertical="top", wrap_text=True)


def sheet_name_for(index: int, label: str, used: set) -> str:
    slug = INVALID_SHEET_CHARS_RE.sub(" ", label)
    slug = WHITESPACE_RE.sub(" ", slug).strip()
    base = "Q{} {}".format(index, slug).strip()[:MAX_SHEET_NAME].strip()
    name = base or "Q{}".format(index)
    suffix = 2
    while name.lower() in used:
        trimmed = base[: MAX_SHEET_NAME - len(str(suffix)) - 1].strip()
        name = "{} {}".format(trimmed, suffix)
        suffix += 1
    used.add(name.lower())
    return name


def write_header_row(sheet, row: int, labels: Sequence[str]) -> None:
    for column, label in enumerate(labels, start=1):
        cell = sheet.cell(row=row, column=column, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")


def set_widths(sheet, widths: Sequence[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def write_summary_sheet(sheet, groups: Sequence[QuestionGroup], file_reports: Sequence[Dict[str, object]], sheet_names: Sequence[str]) -> None:
    sheet.title = "Summary"
    sheet["A1"] = "Transcript responses by question"
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = "Generated {}".format(datetime.now().strftime("%d %b %Y, %I:%M %p"))
    sheet["A3"] = "{} source file(s) · {} question(s) · {} response(s)".format(
        len(file_reports),
        len(groups),
        sum(len(group.responses) for group in groups),
    )

    row = 5
    write_header_row(sheet, row, ["Source file", "Detected layout", "Rows read", "Responses"])
    for report in file_reports:
        row += 1
        sheet.cell(row=row, column=1, value=report["file"]).alignment = TOP_WRAP
        sheet.cell(row=row, column=2, value=report["layout"])
        sheet.cell(row=row, column=3, value=report["rows"])
        sheet.cell(row=row, column=4, value=report["responses"])

    row += 2
    write_header_row(sheet, row, ["#", "Question", "Responses", "Transcripts", "Blank", "Sheet"])
    for index, (group, name) in enumerate(zip(groups, sheet_names), start=1):
        row += 1
        sheet.cell(row=row, column=1, value=index)
        label_cell = sheet.cell(row=row, column=2, value=truncate(group.label))
        label_cell.alignment = TOP_WRAP
        sheet.cell(row=row, column=3, value=len(group.responses))
        sheet.cell(row=row, column=4, value=len(group.transcripts))
        sheet.cell(row=row, column=5, value=group.unanswered)
        sheet.cell(row=row, column=6, value=name)

    set_widths(sheet, [8, 70, 12, 12, 10, 24])


def write_combined_sheet(sheet, groups: Sequence[QuestionGroup]) -> None:
    sheet.title = "All responses"
    row = 1
    for index, group in enumerate(groups, start=1):
        question_cell = sheet.cell(
            row=row, column=1, value="Q{}. {}".format(index, truncate(group.label))
        )
        question_cell.font = QUESTION_FONT
        question_cell.alignment = TOP_WRAP
        for column in range(1, 4):
            sheet.cell(row=row, column=column).fill = QUESTION_FILL
        sheet.row_dimensions[row].height = 28
        row += 1

        write_header_row(sheet, row, ["Transcript", "Speaker", "Response"])
        row += 1

        if not group.responses:
            cell = sheet.cell(row=row, column=1, value="No responses captured")
            cell.font = Font(italic=True, color="6B7280")
            row += 2
            continue

        for response in group.responses:
            sheet.cell(row=row, column=1, value=response.transcript).alignment = TOP_WRAP
            sheet.cell(row=row, column=2, value=response.speaker or "—").alignment = TOP_WRAP
            sheet.cell(row=row, column=3, value=truncate(response.answer)).alignment = TOP_WRAP
            row += 1
        row += 1

    set_widths(sheet, [34, 22, 110])
    sheet.freeze_panes = "A1"


def write_question_sheet(sheet, index: int, group: QuestionGroup) -> None:
    sheet["A1"] = "Q{}".format(index)
    sheet["A1"].font = HEADER_FONT
    sheet["B1"] = truncate(group.label)
    sheet["B1"].font = QUESTION_FONT
    sheet["B1"].alignment = TOP_WRAP
    sheet["A2"] = "Responses"
    sheet["A2"].font = HEADER_FONT
    sheet["B2"] = "{} across {} transcript(s)".format(
        len(group.responses), len(group.transcripts)
    )

    write_header_row(sheet, 4, ["Transcript", "Speaker", "Response", "Source file"])
    row = 5
    for response in group.responses:
        sheet.cell(row=row, column=1, value=response.transcript).alignment = TOP_WRAP
        sheet.cell(row=row, column=2, value=response.speaker or "—").alignment = TOP_WRAP
        sheet.cell(row=row, column=3, value=truncate(response.answer)).alignment = TOP_WRAP
        sheet.cell(row=row, column=4, value=response.source_file).alignment = TOP_WRAP
        row += 1

    if not group.responses:
        cell = sheet.cell(row=row, column=1, value="No responses captured")
        cell.font = Font(italic=True, color="6B7280")

    set_widths(sheet, [34, 22, 110, 30])
    sheet.freeze_panes = "A5"


def write_workbook(
    output_path: str,
    groups: Sequence[QuestionGroup],
    file_reports: Sequence[Dict[str, object]],
    per_question_sheets: bool = True,
) -> List[str]:
    workbook = Workbook()
    used_names = {"summary", "all responses"}
    sheet_names: List[str] = []
    if per_question_sheets:
        for index, group in enumerate(groups, start=1):
            sheet_names.append(sheet_name_for(index, group.label, used_names))
    else:
        sheet_names = ["All responses"] * len(groups)

    write_summary_sheet(workbook.active, groups, file_reports, sheet_names)
    write_combined_sheet(workbook.create_sheet(), groups)
    if per_question_sheets:
        for index, (group, name) in enumerate(zip(groups, sheet_names), start=1):
            write_question_sheet(workbook.create_sheet(title=name), index, group)

    directory = os.path.dirname(os.path.abspath(output_path))
    if directory:
        os.makedirs(directory, exist_ok=True)
    workbook.save(output_path)
    return sheet_names


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Consolidate transcript CSVs into an Excel workbook grouped by question."
    )
    parser.add_argument("inputs", nargs="+", help="One or more transcript CSV files")
    parser.add_argument(
        "-o",
        "--output",
        default="transcript-responses.xlsx",
        help="Path of the .xlsx file to write",
    )
    parser.add_argument(
        "--no-question-sheets",
        action="store_true",
        help="Only write Summary and All responses sheets",
    )
    parser.add_argument(
        "--print-summary",
        action="store_true",
        help="Print a JSON summary of the result to stdout",
    )
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)

    missing = [path for path in args.inputs if not os.path.isfile(path)]
    if missing:
        sys.stderr.write("File not found: {}\n".format(", ".join(missing)))
        return 1

    groups, file_reports = build_groups(args.inputs)
    if not groups:
        sys.stderr.write(
            "No questions found. Expected a question/answer, speaker/text, "
            "or one-column-per-question CSV.\n"
        )
        return 1

    ordered = list(groups.values())
    sheet_names = write_workbook(
        args.output, ordered, file_reports, per_question_sheets=not args.no_question_sheets
    )

    summary = {
        "output": os.path.abspath(args.output),
        "files": file_reports,
        "questionCount": len(ordered),
        "responseCount": sum(len(group.responses) for group in ordered),
        "transcriptCount": len(
            {response.transcript for group in ordered for response in group.responses}
        ),
        "questions": [
            {
                "index": index,
                "question": group.label,
                "sheet": name,
                "responseCount": len(group.responses),
                "transcriptCount": len(group.transcripts),
                "blankCount": group.unanswered,
                "sampleResponses": [
                    {
                        "transcript": response.transcript,
                        "speaker": response.speaker,
                        "answer": truncate(response.answer, 400),
                    }
                    for response in group.responses[:3]
                ],
            }
            for index, (group, name) in enumerate(zip(ordered, sheet_names), start=1)
        ],
    }

    if args.print_summary:
        sys.stdout.write(json.dumps(summary))
    else:
        sys.stdout.write(
            "Wrote {} — {} question(s), {} response(s) from {} file(s)\n".format(
                summary["output"],
                summary["questionCount"],
                summary["responseCount"],
                len(file_reports),
            )
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
